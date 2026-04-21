"""Dump structure + sample data from the Division 1 fixtures xlsx.

Usage:
    py _fixtures.py > fixture_dump.txt
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
KNOWLEDGE = HERE.parent

DIV1 = KNOWLEDGE / "National ESCOCCER LEAGUE- DIVISION 1 .xlsx"
DIV2 = KNOWLEDGE / "EVENT FLOW DIVISION 2 CADE ESPORTS ESOCCER.xlsx"


def dump(path: Path, label: str) -> None:
    print("=" * 80)
    print(f"FILE: {label}")
    print(f"PATH: {path}")
    print("=" * 80)
    if not path.exists():
        print(f"MISSING FILE: {path}")
        return
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    print(f"Sheet names: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        print()
        print("-" * 80)
        print(f"SHEET: {name!r}")
        print(f"  dims: max_row={ws.max_row} max_col={ws.max_column}")
        print("-" * 80)
        # Dump ALL rows (fixtures workbooks aren't huge)
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Trim trailing Nones
            trimmed = list(row)
            while trimmed and trimmed[-1] in (None, ""):
                trimmed.pop()
            if not trimmed:
                print(f"  R{row_idx:>3}: (empty)")
                continue
            cells = " | ".join(
                "" if v is None else str(v).replace("\n", "\\n")
                for v in trimmed
            )
            print(f"  R{row_idx:>3}: {cells}")


def main() -> int:
    dump(DIV1, "Division 1 fixtures")
    print()
    print()
    dump(DIV2, "Division 2 event flow (context only)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
