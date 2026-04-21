"""
Plan 18 — extract real H2H fixture pairings from the Division 1 event-flow xlsx.

The xlsx has one sheet per match day (WK N DAY N). Each H2H row carries the
Division 1 matchup and the Division 2 matchup side-by-side in adjacent cells.
This script reads only the Div 1 column, normalizes gamer tags against the
13-player Elite roster, and emits a per-match-day structured dump.

Usage:
    py _real_fixtures.py > real_fixtures.txt
"""

from __future__ import annotations
import re
import sys
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent.parent  # ESOCCER/
XLSX = ROOT / "KNOWLEDGE" / "EVENT FLOW DIVISION 2 CADE ESPORTS ESOCCER.xlsx"

ROSTER = {
    "adefola": "ADEFOLA",
    "anife": "ANIFE",
    "baji jnr": "BAJI_JNR",
    "baji junior": "BAJI_JNR",
    "baji j": "BAJI_JNR",
    "baji snr": "BAJI_JNR",   # xlsx uses SNR — treat same person for Plan 18
    "baji senior": "BAJI_JNR",
    "dadaboi": "DADABOI",
    "faruk": "FARUK",
    "guru": "GURU",
    "kaykay": "KAYKAY",
    "kay kay": "KAYKAY",
    "killer freak": "KILLER_FREAK",
    "killerfreak": "KILLER_FREAK",
    "kingnonex": "KINGNONEX",
    "king nonex": "KINGNONEX",
    "mitch": "MITCH",
    "mr oga": "MR_OGA",
    "mr. oga": "MR_OGA",
    "oga": "MR_OGA",
    "tactical": "TACTICAL",
    "wolevation": "WOLEVATION",
}


def normalize(name: str) -> str | None:
    n = re.sub(r"[^a-z0-9 ]", "", name.strip().lower())
    n = re.sub(r"\s+", " ", n).strip()
    if not n:
        return None
    if n in ROSTER:
        return ROSTER[n]
    # Try prefix match for odd variants ("KAYKAY JNR", etc.)
    for k, v in ROSTER.items():
        if n.startswith(k) or k.startswith(n):
            return v
    return None  # not in our 13


VS = re.compile(r"^(.+?)\s+vs?\.?\s+(.+?)$", re.IGNORECASE)


def parse_vs(text: str):
    m = VS.match(text.strip())
    if not m:
        return None
    left, right = m.group(1), m.group(2)
    return (left.strip(), right.strip())


def find_day_sheets(wb):
    return [s for s in wb.sheetnames if re.match(r"^WK\s*\d+\s*DAY\s*\d+\s*$", s, re.IGNORECASE)]


def iter_vs_cells(ws):
    """Yield (row, tuple) for every cell in the sheet that looks like 'X vs Y'."""
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if not isinstance(cell, str):
                continue
            parts = parse_vs(cell)
            if parts:
                yield row, parts


def main() -> int:
    if not XLSX.exists():
        print(f"Missing {XLSX}", file=sys.stderr)
        return 1
    wb = load_workbook(XLSX, data_only=True)
    day_sheets = find_day_sheets(wb)
    if not day_sheets:
        print("No WK N DAY N sheets found.", file=sys.stderr)
        return 1

    # Assign monotonic match_date — Saturday weekly from 2025-11-01.
    from datetime import timedelta
    anchor = date(2025, 11, 1)
    per_day_offset = [0, 1]  # WK 1 DAY 1 = Sat, WK 1 DAY 2 = Sun

    md_count = 0
    match_total = 0
    out = []
    for idx, sheet_name in enumerate(day_sheets):
        m = re.match(r"WK\s*(\d+)\s*DAY\s*(\d+)", sheet_name, re.IGNORECASE)
        week = int(m.group(1))
        day = int(m.group(2))
        # week-1 gives 0-indexed week offset; day 1 => Saturday, day 2 => Sunday
        md_date = anchor + timedelta(weeks=(week - 1), days=(day - 1))
        ws = wb[sheet_name]

        pairings = []
        seen = set()
        for row, parts in iter_vs_cells(ws):
            left_raw, right_raw = parts
            # Skip "Rule 3.4.4.2: ... vs ..." style phrase matches by length
            if len(left_raw) > 40 or len(right_raw) > 40:
                continue
            left = normalize(left_raw)
            right = normalize(right_raw)
            if not left or not right:
                continue
            if left == right:
                continue
            key = tuple(sorted((left, right)))
            if key in seen:
                continue
            seen.add(key)
            pairings.append({
                "home_raw": left_raw,
                "away_raw": right_raw,
                "home": left,
                "away": right,
            })

        if not pairings:
            continue
        md_count += 1
        match_total += len(pairings)
        out.append({
            "sheet": sheet_name,
            "week": week,
            "day": day,
            "match_date": md_date.isoformat(),
            "pairings": pairings,
        })

    print(f"# Real fixture pairings extracted from {XLSX.name}")
    print(f"# match_days: {md_count}")
    print(f"# matches   : {match_total}")
    print()

    unmatched_names: set[str] = set()
    # second pass to log unrecognized names
    for sheet_name in day_sheets:
        ws = wb[sheet_name]
        for _, (l, r) in ((row, parts) for row, parts in iter_vs_cells(ws)):
            if len(l) > 40 or len(r) > 40:
                continue
            if not normalize(l):
                unmatched_names.add(l.strip())
            if not normalize(r):
                unmatched_names.add(r.strip())

    for md in out:
        print(f"## {md['sheet']} — {md['match_date']}")
        for p in md["pairings"]:
            print(f"  {p['home']} vs {p['away']}  ({p['home_raw']} vs {p['away_raw']})")
        print()

    if unmatched_names:
        print("\n## Names not in 13-player Elite roster (likely Div 2 or different Elite cohort):")
        for n in sorted(unmatched_names):
            print(f"  - {n}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
